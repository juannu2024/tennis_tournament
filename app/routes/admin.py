from flask import Blueprint, render_template, redirect, url_for, request, flash, send_file
from flask_login import login_required
from flask_wtf import FlaskForm
from flask_wtf.file import FileField, FileAllowed
from wtforms import StringField, DateTimeField, SelectField, TextAreaField
from wtforms.validators import DataRequired
from ..models.models import db, Tournament, Player
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tempfile
import os
from werkzeug.utils import secure_filename

bp = Blueprint('admin', __name__, url_prefix='/admin')

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'tournament_images')

# Asegurarse de que el directorio de subida existe
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

class TournamentForm(FlaskForm):
    name = StringField('Nombre', validators=[DataRequired()])
    start_date = DateTimeField('Fecha de inicio', format='%Y-%m-%d', validators=[DataRequired()])
    end_date = DateTimeField('Fecha de fin', format='%Y-%m-%d', validators=[DataRequired()])
    status = SelectField('Estado', choices=[
        ('active', 'Activo'),
        ('inactive', 'Inactivo'),
        ('upcoming', 'Próximo'),
        ('finished', 'Terminado')
    ], validators=[DataRequired()])
    image = FileField('Cartel del Torneo', validators=[
        FileAllowed(ALLOWED_EXTENSIONS, 'Solo se permiten imágenes')
    ])

def save_tournament_image(form_image):
    if form_image:
        filename = secure_filename(form_image.filename)
        # Añadir timestamp al nombre para evitar duplicados
        name, ext = os.path.splitext(filename)
        filename = f"{name}_{int(datetime.utcnow().timestamp())}{ext}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        form_image.save(filepath)
        return f"tournament_images/{filename}"
    return None

@bp.route('/')
@login_required
def index():
    tournaments = Tournament.query.order_by(Tournament.created_at.desc()).all()
    return render_template('admin/index.html', tournaments=tournaments)

@bp.route('/tournament/new', methods=['GET', 'POST'])
@login_required
def new_tournament():
    form = TournamentForm()
    if form.validate_on_submit():
        if form.status.data == 'active':
            # Desactivar cualquier torneo activo actual
            active_tournament = Tournament.query.filter_by(status='active').first()
            if active_tournament:
                active_tournament.status = 'inactive'
                db.session.commit()

        image_path = save_tournament_image(form.image.data)
        
        tournament = Tournament(
            name=form.name.data,
            start_date=form.start_date.data,
            end_date=form.end_date.data,
            status=form.status.data,
            image=image_path
        )
        db.session.add(tournament)
        db.session.commit()
        flash('Torneo creado exitosamente.', 'success')
        return redirect(url_for('admin.index'))
    return render_template('admin/tournament_form.html', form=form, tournament=None)

@bp.route('/tournament/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_tournament(id):
    tournament = Tournament.query.get_or_404(id)
    form = TournamentForm(obj=tournament)
    if form.validate_on_submit():
        if form.status.data == 'active':
            # Desactivar cualquier otro torneo activo
            active_tournament = Tournament.query.filter(
                Tournament.status == 'active',
                Tournament.id != id
            ).first()
            if active_tournament:
                active_tournament.status = 'inactive'
                db.session.commit()

        # Manejar la imagen
        if form.image.data:
            # Si hay una imagen anterior, eliminarla
            if tournament.image:
                old_image_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 
                                            'static', tournament.image)
                if os.path.exists(old_image_path):
                    os.remove(old_image_path)
            
            # Guardar la nueva imagen
            image_path = save_tournament_image(form.image.data)
            tournament.image = image_path

        tournament.name = form.name.data
        tournament.start_date = form.start_date.data
        tournament.end_date = form.end_date.data
        tournament.status = form.status.data
        
        db.session.commit()
        flash('Torneo actualizado exitosamente.', 'success')
        return redirect(url_for('admin.index'))
    
    return render_template('admin/tournament_form.html', form=form, tournament=tournament)

@bp.route('/tournament/<int:id>/delete', methods=['POST'])
@login_required
def delete_tournament(id):
    tournament = Tournament.query.get_or_404(id)
    db.session.delete(tournament)
    db.session.commit()
    flash('Torneo eliminado con éxito', 'success')
    return redirect(url_for('admin.index'))

@bp.route('/tournament/<int:id>/players')
@login_required
def tournament_players(id):
    tournament = Tournament.query.get_or_404(id)
    players = Player.query.filter_by(tournament_id=id).order_by(Player.category).all()
    return render_template('admin/players.html', tournament=tournament, players=players)

@bp.route('/player/<int:id>/delete', methods=['POST'])
@login_required
def delete_player(id):
    player = Player.query.get_or_404(id)
    tournament_id = player.tournament_id
    db.session.delete(player)
    db.session.commit()
    flash('Jugador eliminado con éxito', 'success')
    return redirect(url_for('admin.tournament_players', id=tournament_id))

@bp.route('/player/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_player(id):
    player = Player.query.get_or_404(id)
    tournament = Tournament.query.get_or_404(player.tournament_id)
    
    if request.method == 'POST':
        player.first_name = request.form['first_name']
        player.last_name = request.form['last_name']
        player.email = request.form['email']
        player.phone = request.form['phone']
        player.license_number = request.form['license_number']
        player.category = request.form['category']
        player.notes = request.form['notes']
        
        try:
            db.session.commit()
            flash('Jugador actualizado correctamente', 'success')
            return redirect(url_for('admin.tournament_players', id=tournament.id))
        except Exception as e:
            db.session.rollback()
            flash('Error al actualizar el jugador', 'danger')
            
    return render_template('admin/edit_player.html', player=player, tournament=tournament)

@bp.route('/tournament/<int:id>/export')
@login_required
def export_players(id):
    tournament = Tournament.query.get_or_404(id)
    players = Player.query.filter_by(tournament_id=id).all()
    
    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jugadores"

    # Definir categorías y sus nombres en español
    categories = {
        'benjamin_male': 'Benjamín Masculino',
        'benjamin_female': 'Benjamín Femenino',
        'alevin_male': 'Alevín Masculino',
        'alevin_female': 'Alevín Femenino',
        'infantil_male': 'Infantil Masculino',
        'infantil_female': 'Infantil Femenino',
        'cadete_male': 'Cadete Masculino',
        'cadete_female': 'Cadete Femenino',
        'absolute_male': 'Absoluto Masculino',
        'absolute_female': 'Absoluto Femenino'
    }

    # Definir estilos
    header_font = Font(bold=True, size=12, color='FFFFFF')
    category_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    category_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    centered = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Configurar ancho de columnas
    ws.column_dimensions['A'].width = 15  # Nombre
    ws.column_dimensions['B'].width = 20  # Apellidos
    ws.column_dimensions['C'].width = 15  # Licencia
    ws.column_dimensions['D'].width = 30  # Email
    ws.column_dimensions['E'].width = 15  # Teléfono
    ws.column_dimensions['F'].width = 30  # Incidencias

    # Título del torneo
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Listado de Jugadores - {tournament.name}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = centered
    title_cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    title_cell.font = Font(bold=True, size=14, color='FFFFFF')

    # Iniciar desde la fila 3
    current_row = 3

    # Procesar cada categoría
    for category_key, category_name in categories.items():
        category_players = [p for p in players if p.category == category_key]
        
        if category_players:
            # Encabezado de categoría
            ws.merge_cells(f'A{current_row}:F{current_row}')
            category_header = ws[f'A{current_row}']
            category_header.value = f"{category_name} ({len(category_players)} jugadores)"
            category_header.font = category_font
            category_header.fill = category_fill
            category_header.alignment = centered
            
            current_row += 1

            # Encabezados de columnas
            headers = ['Nombre', 'Apellidos', 'Licencia', 'Email', 'Teléfono', 'Incidencias']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered
                cell.border = border

            current_row += 1

            # Datos de jugadores
            for player in category_players:
                ws.cell(row=current_row, column=1, value=player.first_name).border = border
                ws.cell(row=current_row, column=2, value=player.last_name).border = border
                ws.cell(row=current_row, column=3, value=player.license_number).border = border
                ws.cell(row=current_row, column=4, value=player.email).border = border
                ws.cell(row=current_row, column=5, value=player.phone).border = border
                ws.cell(row=current_row, column=6, value=player.notes).border = border
                current_row += 1

            # Espacio entre categorías
            current_row += 1

    # Guardar el archivo temporalmente
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    
    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name=f'jugadores_{tournament.name}_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
